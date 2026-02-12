"""Shared export utilities for CSV and Excel report generation."""
from __future__ import annotations

from typing import List, Dict, Any
from io import BytesIO, StringIO
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_csv(data: List[Dict[str, Any]], columns: List[str]) -> str:
    """Export data to CSV string using Python csv module.

    Produces RFC 4180 compliant output with proper quoting and escaping.

    Args:
        data: List of dictionaries with row data
        columns: Ordered list of column names to export

    Returns:
        CSV string with headers and data rows
    """
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')

    writer.writeheader()
    for row in data:
        writer.writerow(row)

    return output.getvalue()


def export_to_excel(
    sheets: Dict[str, Dict[str, Any]],
    title: str = "Regulatory Report",
) -> BytesIO:
    """Export data to Excel workbook with formatting.

    Args:
        sheets: Dict of sheet_name -> {headers: List[str], data: List[Dict], formats: Dict}
        title: Report title for metadata

    Returns:
        BytesIO buffer containing Excel workbook

    Example:
        sheets = {
            "Summary": {
                "headers": ["Metric", "Value"],
                "data": [{"Metric": "Total Allowance", "Value": 1000000}],
                "formats": {"Value": "$#,##0.00"},
            },
        }
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for sheet_name, sheet_config in sheets.items():
        ws = wb.create_sheet(sheet_name)
        headers = sheet_config["headers"]
        data = sheet_config.get("data", [])
        formats = sheet_config.get("formats", {})

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(header)
                cell.value = value
                cell.border = border

                # Apply number format if specified
                if header in formats:
                    cell.number_format = formats[header]

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(header))
            for row_data in data:
                value_length = len(str(row_data.get(header, "")))
                max_length = max(max_length, value_length)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Write to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
